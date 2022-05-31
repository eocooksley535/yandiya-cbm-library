"""
Author: Elvis Obero-Atkins
Last Edited by: Elvis Obero-Atkins

This python script was created to test the viability of using excel instead of CSVs (at least for the MVP)

This script allows the user to in put the amount of columns they would like to view by typing the letters of the columns.

"""

import openpyxl

yandiya_db = openpyxl.load_workbook(
    '.excel_db_test\yandiya-db.xlsx')
all_sheets = yandiya_db.sheetnames

# Default parameters for product-details.xlsx is "ABCDEFGHIJKLMNOPQ"

parameters = str(input(
    "Please enter the the columns in alphabetical order that you wish to display "))

print("All sheet names {} " .format(yandiya_db.sheetnames))


for sheet in all_sheets:
    print("Current sheet name is {}" .format(sheet))
    currentSheet = yandiya_db[sheet]

    for row in range(1, currentSheet.max_row + 1):
        for column in parameters:
            cell_name = "{}{}".format(column, row)
            print("cell position {} has value {}".format(
                cell_name, currentSheet[cell_name].value))
